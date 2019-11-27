# eventbritexl.py - interact with eventbrite data in excel spreadsheets
from openpyxl import Workbook

def save_attendees(attendees:list, filename:str):
    wb = Workbook()
    wp = wb.active

    row = 1
    col = 1
    
    # get questions
    ids = []
    profile_opts = []
    first_att = attendees[len(attendees) - 1]
    for key in first_att["profile"].keys():
        profile_opts.append(key)
        wp.cell(row = row, column=col).value = key
        col += 1
    for question in first_att["answers"]:
        head = question["question"] + ":" + question["question_id"]
        wp.cell(row=row, column=col).value = head
        ids.append(question["question_id"])
        col += 1
    col = 1
    row += 1

    # save responses
    for attendee in attendees:
        for opt in profile_opts:
            try:
                wp.cell(row=row, column=col).value = str(attendee["profile"][opt])
            except KeyError:
                wp.cell(row=row, column=col).value = "_"
            col += 1
        i = 0
        for question in attendee["answers"]:
            try:
                if question["question_id"] != ids[i]:
                    wp.cell(row=row, column=col).value = question["answer"]
                    i += 1
                else:
                     wp.cell(row=row, column=col).value = "_"
            except KeyError:
                wp.cell(row=row, column=col).value = "_"
            col += 1
        col = 1
        row += 1
    wb.save(filename)
